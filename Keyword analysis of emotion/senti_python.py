import jieba
import numpy as np
import openpyxl as xl

def open_dict(Dict='hahah', path=r'/Users/PycharmProjects/Textming/Sent_Dict/Hownet/'):
    path = path + '%s.txt' % Dict
    dictionary = open(path, 'r', encoding='utf-8')
    dict = []
    for word in dictionary:
        word = word.strip('\n')
        dict.append(word)
    return dict


def judgeodd(num):
    if (num % 2) == 0:
        return 'even'
    else:
        return 'odd'

    
deny_word = open_dict(Dict='not', path=r'/Users/admin/PycharmProjects/HelloWorld/Textming/')
posdict = open_dict(Dict='positive', path=r'/Users/admin/PycharmProjects/HelloWorld/Textming/')
negdict = open_dict(Dict='negative', path=r'/Users/admin/PycharmProjects/HelloWorld/Textming/')

degree_word = open_dict(Dict='Degree level', path=r'/Users/admin/PycharmProjects/HelloWorld/Textming/')
mostdict = degree_word[degree_word.index('extreme') + 1: degree_word.index('very')]  
verydict = degree_word[degree_word.index('very') + 1: degree_word.index('more')]  
moredict = degree_word[degree_word.index('more') + 1: degree_word.index('ish')]  
ishdict = degree_word[degree_word.index('ish') + 1: degree_word.index('last')]  


def sentiment_score_list(dataset):
    seg_sentence = dataset.split('。')
    count1 = []
    count2 = []
    print(seg_sentence)
    for sen in seg_sentence:  
        segtmp = jieba.lcut(sen, cut_all=False) 
        for i in segtmp:
            if ' ' in segtmp:
                segtmp.remove(' ')
        i = 0  
        a = 0  
        poscount = 0  
        poscount2 = 0  
        poscount3 = 0  
        negcount = 0
        negcount2 = 0
        negcount3 = 0
        for word in segtmp:
            if word in posdict:  
                poscount += 1
                c = 0
                for w in segtmp[a:i]:  
                    if w in mostdict:
                        poscount *= 2.0
                    elif w in verydict:
                        poscount *= 1.5
                    elif w in moredict:
                        poscount *= 1.0
                    elif w in ishdict:
                        poscount *= 0.5
                    elif w in deny_word:
                        c += 1
                if judgeodd(c) == 'odd':  
                    poscount *= -1.0
                    poscount2 += poscount
                    poscount = 0
                    poscount3 = poscount + poscount2 + poscount3
                    poscount2 = 0
                else:
                    poscount3 = poscount + poscount2 + poscount3
                    poscount = 0
                a = i + 1  

            elif word in negdict:  
                negcount += 1
                d = 0
                for w in segtmp[a:i]:
                    if w in mostdict:
                        negcount *= 2.0
                    elif w in verydict:
                        negcount *= 1.5
                    elif w in moredict:
                        negcount *= 1.0
                    elif w in ishdict:
                        negcount *= 0.5
                    elif w in degree_word:
                        d += 1
                if judgeodd(d) == 'odd':
                    negcount *= -1.0
                    negcount2 += negcount
                    negcount = 0
                    negcount3 = negcount + negcount2 + negcount3
                    negcount2 = 0
                else:
                    negcount3 = negcount + negcount2 + negcount3
                    negcount = 0
                a = i + 1
            elif word == '！' or word == '!':  
                for w2 in segtmp[::-1]:  
                    if w2 in posdict or negdict:
                        if poscount3 > negcount3:
                            poscount3 += 2
                        elif poscount3 < negcount3:
                            negcount3 += 2
                        break
            i += 1  

            pos_count = 0
            neg_count = 0
            if poscount3 < 0 and negcount3 > 0:
                neg_count += negcount3 - poscount3
                pos_count = 0
            elif negcount3 < 0 and poscount3 > 0:
                pos_count = poscount3 - negcount3
                neg_count = 0
            elif poscount3 < 0 and negcount3 < 0:
                neg_count = -poscount3
                pos_count = -negcount3
            else:
                pos_count = poscount3
                neg_count = negcount3

        count1.append([pos_count, neg_count])
        count2.append(count1)
        count1 = []
    return count2


def sentiment_score(senti_score_list):
    score = []
    for review in senti_score_list:
        score_array = np.array(review)
        Pos = np.sum(score_array[:, 0])
        Neg = np.sum(score_array[:, 1])
        Pos = Pos.tolist()
        Neg = Neg.tolist()
        score.append([Pos, Neg])
    return score


wb = xl.load_workbook('hhh.xlsx')
sheet = wb['Sheet1']
file_handle = open('1.txt', mode='w')
data = open("C:/Users/admin/PycharmProjects/HelloWorld/Textming/1.txt", "a")

for row in range(2, sheet.max_row + 1):
    cell1 = sheet.cell(row, 1)
    score_value = sentiment_score(sentiment_score_list(cell1.value))
    score_value_new = num_list_new = [str(x) for x in score_value]
    score_value_str = ",".join(score_value_new)
    data.write(score_value_str)
    data.write('\n')
